from openpyxl import Workbook
import pymysql

def Create_Cwa_excel_table(Sclass,Cname,Result):
    stat = True
    CWA_list = []
    work_book = Workbook()
    work_book_sheet = work_book.active
    work_book_sheet.title = ('%s考勤情况' % (Sclass))
    # print(work_book_sheet.title)
    title =['学号','科目','班级','是否出勤','节次','上课日期','时间戳']
    print(Result)
    # 标题
    for j in range(len(title)):
        work_book_sheet.cell(1, column=j+1, value=str(title[j]))
    # 已考勤
    i =0
    cur = 0
    row_list =[]
    for row in Result:
        i = i+1
        CWA_list.append(str(row[0]))
        row_list = list(row)
        del row_list[1]
        row = tuple(row_list)
        for m in range(len(row)):
            work_book_sheet.cell(row=i+1, column=m+1, value=str(row[m]))
            cur = i+1
            print(cur)
    # 未考勤
    stu_list = Student_Sno_Query(Sclass)
    row1 = Result[0]
    example = [row1[2],row1[3],'否',row1[5],row1[6],row1[7]]
    for k in stu_list:
        if k not in CWA_list:
            for l in range(len(example)):
                work_book_sheet.cell(row=cur+1,column=1,value=str(k))
                print(cur)
                work_book_sheet.cell(row=cur+1,column=l+2,value=str(example[l]))


    work_book.save('./excel/%s%s考勤表.xlsx' % (Sclass,Cname))
    return stat

def Student_Sno_Query(Sclass):
    conn = pymysql.connect(host = '192.168.123.209',user = 'root',passwd = '123456',db = 'kaoqinxitong',charset = 'utf8')
    c = conn.cursor()
    Sno = []
    sql_str = "SELECT Sno FROM Student WHERE Sclass ='"+Sclass+"'"
    try:
        c.execute(sql_str)
        # conn.commit()
        # print (c.fetchall())
        result = c.fetchall()
        for row in result:
            Sno.append(row[0])
    except:
        conn.rollback()
    conn.close()
    return Sno
# Result = [('0914150230', '关羽', '数学', '09141502', '是', '1-2节', '2019/4/21', '2019-04-21 13:02:49')]
# Create_Cwa_excel_table("asd",Result)